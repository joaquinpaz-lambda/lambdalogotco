#!/usr/bin/env python3
"""
ML Hires Signal Agent
Weekly pipeline: queries TheirStack + Adzuna for ML/AI job postings at 141 enterprise companies,
computes week-over-week deltas, scores accounts into heat tiers, outputs Excel heatmap.
"""

import json
import os
import sys
import time
import logging
from datetime import datetime, timedelta
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).parent
COMPANIES_FILE = BASE_DIR / "companies.json"
STATE_FILE = BASE_DIR / "data" / "ml_hires_state.json"
OUTPUT_DIR = BASE_DIR / "output"

THEIRSTACK_API_KEY = os.environ.get("THEIRSTACK_API_KEY", "")
ADZUNA_APP_ID = os.environ.get("ADZUNA_APP_ID", "")
ADZUNA_API_KEY = os.environ.get("ADZUNA_API_KEY", "")

THEIRSTACK_ENDPOINT = "https://api.theirstack.com/v1/jobs/search"
ADZUNA_ENDPOINT = "https://api.adzuna.com/v1/api/jobs/us/search/1"

# ML title keywords
ML_KEYWORDS = [
    "Machine Learning", "ML Engineer", "AI Engineer", "Deep Learning",
    "NLP Engineer", "Computer Vision", "Data Scientist", "Applied Scientist",
    "Research Scientist", "MLOps", "LLM", "Foundation Model", "Generative AI",
    "GenAI", "AI/ML", "ML Platform", "AI Platform", "AI Infrastructure",
    "Inference Engineer", "GPU Engineer", "CUDA", "Model Training",
    "Reinforcement Learning", "Neural Network",
]

# Director+ title patterns
DIRECTOR_KEYWORDS = [
    "Director", "Senior Director", "Sr. Director", "VP", "Vice President",
    "Head of", "Principal", "Distinguished", "Fellow", "Chief AI",
    "Chief Data", "Executive Director",
]

# GPU/Inference signal keywords
GPU_KEYWORDS = [
    "GPU", "CUDA", "inference", "serving", "H100", "A100", "H200", "TPU",
    "accelerator", "training infrastructure", "model serving",
    "distributed training", "vLLM", "TensorRT", "Triton",
]

# Scoring thresholds
HOT_MIN_NET_NEW = 5
HOT_MIN_PCT = 25.0
HOT_DIRECTOR_NET_NEW = 2
WARM_MIN_NET_NEW = 1

# Excel colors
COLOR_HOT_BG = "C6EFCE"       # green fill
COLOR_HOT_FONT = "276221"
COLOR_WARM_BG = "FFEB9C"      # yellow fill
COLOR_WARM_FONT = "9C6500"
COLOR_QUIET_BG = "FFFFFF"     # white fill
COLOR_QUIET_FONT = "000000"
COLOR_HEADER_BG = "1A1A1A"    # Lambda Labs dark header
COLOR_HEADER_FONT = "FFFFFF"

# Lookback window: jobs posted in the last 7 days
LOOKBACK_DAYS = 7

# TheirStack: batch up to 10 domains per request to reduce API calls
THEIRSTACK_BATCH_SIZE = 10

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# State persistence
# ---------------------------------------------------------------------------

def load_state() -> dict:
    """Load last week's state. Returns {} if no prior state."""
    if STATE_FILE.exists():
        with open(STATE_FILE) as f:
            return json.load(f)
    return {}


def save_state(state: dict) -> None:
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2)
    log.info("State saved to %s", STATE_FILE)


# ---------------------------------------------------------------------------
# TheirStack API
# ---------------------------------------------------------------------------

def query_theirstack_batch(domains: list[str], since_date: str) -> dict[str, list[dict]]:
    """
    Query TheirStack for multiple domains in one request.
    Returns dict: domain -> list of job dicts.
    """
    if not THEIRSTACK_API_KEY:
        log.warning("THEIRSTACK_API_KEY not set — skipping TheirStack query")
        return {d: [] for d in domains}

    payload = {
        "page": 0,
        "limit": 100,
        "company_domain_or": domains,
        "job_title_or": ML_KEYWORDS,
        "posted_at_gte": since_date,
        "order_by": [{"desc": True, "field": "date_posted"}],
    }
    headers = {
        "Authorization": f"Bearer {THEIRSTACK_API_KEY}",
        "Content-Type": "application/json",
    }

    results = {d: [] for d in domains}
    page = 0
    total_fetched = 0

    while True:
        payload["page"] = page
        try:
            resp = requests.post(
                THEIRSTACK_ENDPOINT,
                json=payload,
                headers=headers,
                timeout=30,
            )
            if resp.status_code == 429:
                log.warning("TheirStack rate limit hit — sleeping 60s")
                time.sleep(60)
                continue
            if resp.status_code == 402:
                log.error("TheirStack: payment required / quota exceeded")
                break
            resp.raise_for_status()
            data = resp.json()
        except requests.RequestException as exc:
            log.error("TheirStack request failed for domains %s: %s", domains, exc)
            break

        jobs = data.get("data", [])
        if not jobs:
            break

        for job in jobs:
            # Match job back to domain
            company_obj = job.get("company", {}) or {}
            job_domain = (company_obj.get("domain") or "").lower().strip()
            # Try to match against our requested domains
            matched = False
            for d in domains:
                if d.lower() in job_domain or job_domain in d.lower():
                    results[d].append(job)
                    matched = True
                    break
            if not matched:
                # Fallback: check company name against any domain
                pass

        total_fetched += len(jobs)
        page += 1

        # TheirStack paginates — check if we got a full page
        if len(jobs) < payload["limit"]:
            break
        # Safety: max 5 pages per batch
        if page >= 5:
            break

    log.debug("TheirStack batch %s: fetched %d jobs across %d domains",
              domains[:2], total_fetched, len(domains))
    return results


def query_theirstack_all(companies: list[dict], since_date: str) -> dict[str, list[dict]]:
    """Query TheirStack for all companies in batches."""
    all_results: dict[str, list[dict]] = {}
    domains = [c["domain"] for c in companies]

    for i in range(0, len(domains), THEIRSTACK_BATCH_SIZE):
        batch = domains[i: i + THEIRSTACK_BATCH_SIZE]
        log.info(
            "TheirStack batch %d-%d / %d",
            i + 1, min(i + THEIRSTACK_BATCH_SIZE, len(domains)), len(domains),
        )
        batch_results = query_theirstack_batch(batch, since_date)
        all_results.update(batch_results)
        # Polite delay between batches
        if i + THEIRSTACK_BATCH_SIZE < len(domains):
            time.sleep(1)

    return all_results


# ---------------------------------------------------------------------------
# Adzuna API (fallback / supplement)
# ---------------------------------------------------------------------------

def query_adzuna(company_name: str, since_date: str) -> list[dict]:
    """Query Adzuna for a single company. Returns list of job dicts."""
    if not ADZUNA_APP_ID or not ADZUNA_API_KEY:
        return []

    # Build keyword string from ML keywords (top 5 most specific)
    what = "machine learning OR AI engineer OR deep learning OR MLOps OR generative AI"
    params = {
        "app_id": ADZUNA_APP_ID,
        "app_key": ADZUNA_API_KEY,
        "results_per_page": 50,
        "what": what,
        "company": company_name,
        "content-type": "application/json",
        "max_days_old": LOOKBACK_DAYS,
    }
    try:
        resp = requests.get(ADZUNA_ENDPOINT, params=params, timeout=20)
        if resp.status_code == 429:
            log.warning("Adzuna rate limit for %s", company_name)
            time.sleep(30)
            return []
        resp.raise_for_status()
        data = resp.json()
        return data.get("results", [])
    except requests.RequestException as exc:
        log.warning("Adzuna failed for %s: %s", company_name, exc)
        return []


# ---------------------------------------------------------------------------
# Job analysis helpers
# ---------------------------------------------------------------------------

def is_ml_job(title: str) -> bool:
    title_lower = title.lower()
    return any(kw.lower() in title_lower for kw in ML_KEYWORDS)


def is_director_plus(title: str) -> bool:
    title_lower = title.lower()
    return any(kw.lower() in title_lower for kw in DIRECTOR_KEYWORDS)


def has_gpu_signal(title: str, description: str = "") -> bool:
    combined = (title + " " + description).lower()
    return any(kw.lower() in combined for kw in GPU_KEYWORDS)


def extract_top_role(jobs: list[dict], source: str = "theirstack") -> str:
    """Return the title of the highest-signal job."""
    director_jobs = []
    gpu_jobs = []
    other_jobs = []

    for job in jobs:
        if source == "theirstack":
            title = job.get("job_title", "") or ""
            desc = job.get("description", "") or ""
        else:
            title = job.get("title", "") or ""
            desc = job.get("description", "") or ""

        if is_director_plus(title):
            director_jobs.append(title)
        elif has_gpu_signal(title, desc):
            gpu_jobs.append(title)
        else:
            other_jobs.append(title)

    candidates = director_jobs or gpu_jobs or other_jobs
    if not candidates:
        return ""
    # Return first (most recent since API returns ordered by date)
    return candidates[0][:80]


def build_notable_signal(
    net_new: int,
    has_director: bool,
    has_gpu: bool,
    total_this_week: int,
) -> str:
    parts = []
    if has_director:
        parts.append("Director+ hiring")
    if has_gpu:
        parts.append("GPU/inference signal")
    if net_new >= HOT_MIN_NET_NEW:
        parts.append(f"+{net_new} new ML roles WoW")
    elif net_new > 0:
        parts.append(f"+{net_new} ML roles added")
    if not parts and total_this_week > 0:
        parts.append(f"{total_this_week} active ML roles")
    return "; ".join(parts) if parts else "—"


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------

TIER_HOT = "Hot"
TIER_WARM = "Warm"
TIER_QUIET = "Quiet"

TIER_ICON = {
    TIER_HOT: "Hot \U0001f7e9",   # 🟩
    TIER_WARM: "Warm \U0001f7e8",  # 🟨
    TIER_QUIET: "Quiet ⬜",    # ⬜
}

ACTION = {
    TIER_HOT: "Outbound Now — High ML Signal",
    TIER_WARM: "Monitor — Trending Up",
    TIER_QUIET: "Hold",
}


def score_account(
    net_new: int,
    wow_pct: float,
    has_director: bool,
    total_this_week: int,
) -> str:
    if (
        net_new >= HOT_MIN_NET_NEW
        or wow_pct >= HOT_MIN_PCT
        or (has_director and net_new >= HOT_DIRECTOR_NET_NEW)
    ):
        return TIER_HOT
    if net_new >= WARM_MIN_NET_NEW or total_this_week > 0:
        return TIER_WARM
    return TIER_QUIET


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def run_pipeline() -> None:
    today = datetime.utcnow().date()
    week_of = today.strftime("%Y-%m-%d")
    since_date = (today - timedelta(days=LOOKBACK_DAYS)).strftime("%Y-%m-%d")

    log.info("=== ML Hires Signal Agent — Week of %s ===", week_of)
    log.info("Querying jobs posted since %s", since_date)

    # Load companies
    with open(COMPANIES_FILE) as f:
        companies: list[dict] = json.load(f)
    log.info("Loaded %d companies", len(companies))

    # Load prior state
    prior_state = load_state()
    log.info(
        "Prior state loaded: %d companies tracked",
        len(prior_state),
    )

    # ---- Query TheirStack (batch) ----
    log.info("Querying TheirStack API...")
    theirstack_results = query_theirstack_all(companies, since_date)

    # ---- Process each company ----
    rows = []
    new_state: dict[str, dict] = {}

    for company in companies:
        domain = company["domain"]
        name = company["company"]

        log.info("Processing: %s (%s)", name, domain)

        # Gather TheirStack jobs
        ts_jobs = theirstack_results.get(domain, [])

        # Filter to confirmed ML jobs (TheirStack already filtered by title,
        # but double-check to avoid false positives)
        ml_ts_jobs = [
            j for j in ts_jobs
            if is_ml_job(j.get("job_title", "") or "")
        ]

        # Adzuna fallback: use if TheirStack returns 0 and Adzuna creds exist
        az_jobs = []
        if len(ml_ts_jobs) == 0 and (ADZUNA_APP_ID and ADZUNA_API_KEY):
            log.debug("Falling back to Adzuna for %s", name)
            az_raw = query_adzuna(name, since_date)
            az_jobs = [
                j for j in az_raw
                if is_ml_job(j.get("title", "") or "")
            ]
            time.sleep(0.5)  # polite

        # Combine counts
        total_this_week = len(ml_ts_jobs) + len(az_jobs)

        # Prior week count
        prior = prior_state.get(domain, {})
        total_last_week = prior.get("count", 0)

        # Delta
        net_new = total_this_week - total_last_week
        if total_last_week > 0:
            wow_pct = (net_new / total_last_week) * 100.0
        elif total_this_week > 0:
            wow_pct = 100.0
        else:
            wow_pct = 0.0

        # Director+ and GPU flags
        all_jobs = ml_ts_jobs + az_jobs

        def get_title(j, is_ts=True):
            return j.get("job_title" if is_ts else "title", "") or ""

        def get_desc(j, is_ts=True):
            return j.get("description", "") or ""

        director_jobs = [
            j for j in ml_ts_jobs if is_director_plus(get_title(j, True))
        ] + [
            j for j in az_jobs if is_director_plus(get_title(j, False))
        ]
        has_director = len(director_jobs) > 0
        director_count = len(director_jobs)

        gpu_jobs_found = [
            j for j in ml_ts_jobs
            if has_gpu_signal(get_title(j, True), get_desc(j, True))
        ] + [
            j for j in az_jobs
            if has_gpu_signal(get_title(j, False), get_desc(j, False))
        ]
        gpu_signal = len(gpu_jobs_found) > 0

        # Score
        tier = score_account(net_new, wow_pct, has_director, total_this_week)

        # Notable signal text
        notable = build_notable_signal(net_new, has_director, gpu_signal, total_this_week)

        # Top role
        top_role_ts = extract_top_role(ml_ts_jobs, "theirstack")
        top_role_az = extract_top_role(az_jobs, "adzuna")
        top_role = top_role_ts or top_role_az or "—"

        # Sources used
        sources_used = []
        if ml_ts_jobs:
            sources_used.append("TheirStack")
        if az_jobs:
            sources_used.append("Adzuna")
        sources_str = ", ".join(sources_used) if sources_used else "—"

        # High signal flag
        high_signal = "Yes" if tier == TIER_HOT else ""

        # GPU/Inference Intent display
        gpu_display = "Yes" if gpu_signal else ""

        rows.append({
            "tier": tier,
            "tier_order": {TIER_HOT: 0, TIER_WARM: 1, TIER_QUIET: 2}[tier],
            "heat": TIER_ICON[tier],
            "pod": company.get("pod", ""),
            "company": name,
            "domain": domain,
            "action": ACTION[tier],
            "net_new": net_new,
            "wow_pct": wow_pct,
            "notable": notable,
            "total_this_week": total_this_week,
            "total_last_week": total_last_week,
            "high_signal": high_signal,
            "director_count": director_count,
            "gpu_display": gpu_display,
            "top_role": top_role,
            "subsector": company.get("subsector", ""),
            "revenue": company.get("revenue", ""),
            "sources": sources_str,
            "week_of": week_of,
        })

        # Update state
        new_state[domain] = {"count": total_this_week, "week_of": week_of}

    # Sort: tier order ASC, then net_new DESC within tier
    rows.sort(key=lambda r: (r["tier_order"], -r["net_new"]))

    # Assign rank
    for i, row in enumerate(rows, start=1):
        row["rank"] = i

    # ---- Write Excel ----
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_file = OUTPUT_DIR / f"ML_Hires_Signal_{week_of}.xlsx"
    write_excel(rows, output_file, week_of)
    log.info("Excel output: %s", output_file)

    # ---- Persist state ----
    save_state(new_state)

    # ---- Summary ----
    hot_count = sum(1 for r in rows if r["tier"] == TIER_HOT)
    warm_count = sum(1 for r in rows if r["tier"] == TIER_WARM)
    quiet_count = sum(1 for r in rows if r["tier"] == TIER_QUIET)
    log.info(
        "Summary — Hot: %d | Warm: %d | Quiet: %d | Total: %d",
        hot_count, warm_count, quiet_count, len(rows),
    )


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

COLUMNS = [
    ("Heat", 14),
    ("Rank", 6),
    ("Pod", 36),
    ("Company", 28),
    ("Domain", 26),
    ("Recommended Action", 32),
    ("Net New ML Roles WoW", 20),
    ("WoW % Change", 14),
    ("Notable Signal", 40),
    ("ML Roles This Week", 19),
    ("ML Roles Last Week", 19),
    ("High Signal", 11),
    ("Director+ Openings This Week", 26),
    ("GPU/Inference Intent", 20),
    ("Top Role In Flight", 40),
    ("Industry Subsector", 22),
    ("Revenue 2024 ($M)", 18),
    ("Sources", 18),
    ("Week Of", 12),
]

COLUMN_KEYS = [
    "heat", "rank", "pod", "company", "domain", "action",
    "net_new", "wow_pct", "notable", "total_this_week", "total_last_week",
    "high_signal", "director_count", "gpu_display", "top_role",
    "subsector", "revenue", "sources", "week_of",
]


def write_excel(rows: list[dict], path: Path, week_of: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = f"ML Hires {week_of}"

    # ---- Header row ----
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 30

    # ---- Data rows ----
    hot_fill = PatternFill("solid", fgColor=COLOR_HOT_BG)
    warm_fill = PatternFill("solid", fgColor=COLOR_WARM_BG)
    quiet_fill = PatternFill("solid", fgColor=COLOR_QUIET_BG)

    hot_font = Font(color=COLOR_HOT_FONT, size=10)
    warm_font = Font(color=COLOR_WARM_FONT, size=10)
    quiet_font = Font(color=COLOR_QUIET_FONT, size=10)

    fill_map = {
        TIER_HOT: (hot_fill, hot_font),
        TIER_WARM: (warm_fill, warm_font),
        TIER_QUIET: (quiet_fill, quiet_font),
    }

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    # Columns that should be centered
    CENTER_COLS = {0, 1, 6, 7, 11, 12, 13, 16, 18}  # 0-indexed

    for row_idx, row in enumerate(rows, start=2):
        tier = row["tier"]
        fill, font = fill_map[tier]

        # Build value list
        values = [
            row["heat"],
            row["rank"],
            row["pod"],
            row["company"],
            row["domain"],
            row["action"],
            row["net_new"],
            f"{row['wow_pct']:.1f}%" if row["wow_pct"] != 0 else "0.0%",
            row["notable"],
            row["total_this_week"],
            row["total_last_week"],
            row["high_signal"],
            row["director_count"] if row["director_count"] else "",
            row["gpu_display"],
            row["top_role"],
            row["subsector"],
            row["revenue"],
            row["sources"],
            row["week_of"],
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.font = font
            cell.border = thin_border

            if (col_idx - 1) in CENTER_COLS:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        ws.row_dimensions[row_idx].height = 20

    # ---- Freeze header ----
    ws.freeze_panes = "A2"

    # ---- Auto-filter ----
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if not THEIRSTACK_API_KEY and not (ADZUNA_APP_ID and ADZUNA_API_KEY):
        log.error(
            "No API credentials found. Set THEIRSTACK_API_KEY "
            "and/or ADZUNA_APP_ID + ADZUNA_API_KEY environment variables."
        )
        sys.exit(1)

    run_pipeline()
